from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

out = 'test_files/student_import_template.xlsx'
wb = Workbook()
# Instructions sheet
ins = wb.active
ins.title = 'INSTRUCTIONS'
ins['A1'] = 'IMPORTANT - READ BEFORE USING'
ins['A2'] = '1. Delete this "INSTRUCTIONS" sheet before importing the file.'
ins['A3'] = '2. Required columns: Name, Roll_Number, Registration_Number, Class.'
ins['A4'] = '3. Optional columns: Section, Father_Name, Father_CNIC, Gender, Date_of_Birth, Parent_Contact, Address, Admission_Date, Image_Name.'
ins['A5'] = '4. Date format must be DD/MM/YYYY (e.g., 31/12/2015).'
ins['A6'] = '5. If you include image filenames, place images in the ZIP file with filenames matching Image_Name.'
ins['A7'] = '6. Remove this sheet before uploading/importing.'

# Template sheet
ws = wb.create_sheet(title='TEMPLATE')
headers = [
    'Name','Roll_Number','Registration_Number','Class','Section','Father_Name',
    'Father_CNIC','Gender','Date_of_Birth','Parent_Contact','Address','Admission_Date','Image_Name'
]
ws.append(headers)
# Example row (no styling)
example = [
    'Ali Ahmed','101','REG-2025-001','Grade-5','A','Ahmed Khan','12345-1234567-1',
    'Male','22/03/2015','03001234567','123 Main St, Lahore','01/04/2025','ali_ahmed.jpg'
]
ws.append(example)

# Set column widths for readability
for i, h in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

# Create a named date style for DD/MM/YYYY
date_style = NamedStyle(name='ddmmyyyy', number_format='DD/MM/YYYY')
# Apply to Date_of_Birth (col 9) and Admission_Date (col 12) for the example and header is text
ws['I2'].style = date_style
ws['L2'].style = date_style

# Save workbook
wb.save(out)
print('Wrote', out)
