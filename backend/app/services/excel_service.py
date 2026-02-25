"""
Excel service for Student Import / Export.

Handles:
- Sample template generation
- File parsing, validation, normalization
- Export to xlsx
- Error report generation
- Formula stripping for security
"""

import re
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.utils.student_id_utils import generate_imported_student_id, validate_student_id_uniqueness

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

TEMPLATE_COLUMNS = [
    # Required columns (must appear first in template)
    "Name",
    "Roll_Number",
    "Registration_Number",
    "Class",
    # Optional columns
    "Section",
    "Father_Name",
    "Father_CNIC",
    "Gender",
    "Date_of_Birth",
    "Parent_Contact",
    "Address",
    "Admission_Date",
    "Image_Name",
]

# Required columns — must be present in uploaded file (normalized keys)
REQUIRED_COLUMNS = {
    "name",
    "roll_number",
    "registration_number",
    "class",
}

# Optional columns — ignored if not present
OPTIONAL_COLUMNS = {
    "section",
    "father_name",
    "father_cnic",
    "gender",
    "date_of_birth",
    "parent_contact",
    "address",
    "admission_date",
    "image_name",
}

# Column name variations mapping (for fuzzy column matching)
COLUMN_ALIASES = {
    # Name variations
    "name": ["name", "full_name", "student_name", "fullname", "student"],
    "roll_number": ["roll_number", "roll_no", "rollno", "roll", "rollnumber"],
    "registration_number": ["registration_number", "reg_number", "reg_no", "regno", "registration", "reg", "student_id", "studentid", "id"],
    "class": ["class", "class_name", "classname", "grade", "standard", "class_id"],
    "section": ["section", "sec", "division", "div"],
    "father_name": ["father_name", "fathername", "father", "parent_name", "parentname", "guardian_name", "guardianname"],
    "father_cnic": ["father_cnic", "parent_cnic", "cnic", "b_form", "bform", "nic"],
    "gender": ["gender", "sex"],
    "date_of_birth": ["date_of_birth", "dob", "dateofbirth", "birth_date", "birthdate"],
    "parent_contact": ["parent_contact", "parentcontact", "phone", "contact", "mobile", "guardian_contact", "guardiancontact", "father_phone", "fatherphone"],
    "address": ["address", "home_address", "homeaddress", "residence"],
    "admission_date": ["admission_date", "admissiondate", "joining_date", "joiningdate", "enrolled_date"],
    "image_name": ["image_name", "imagename", "photo", "picture", "image", "photo_name", "photoname"],
}

EXAMPLE_ROW = [
    "Ali Ahmed",
    "101",
    "REG-2025-001",
    "Grade-5",
    "A",
    "Ahmed Khan",
    "12345-1234567-1",
    "Male",
    "22/03/2015",
    "03001234567",
    "123 Main St, Lahore",
    "01/04/2025",
    "ali_ahmed.jpg",
]

GENDER_MAP: Dict[str, str] = {
    "m": "Male",
    "male": "Male",
    "boy": "Male",
    "f": "Female",
    "female": "Female",
    "girl": "Female",
    "o": "Other",
    "other": "Other",
}

# ---------------------------------------------------------------------------
# Security — strip Excel injection formulas
# ---------------------------------------------------------------------------

_FORMULA_RE = re.compile(r"^[\s]*[=+\-@]")


def _sanitize_cell(value: Any) -> Any:
    """Strip leading formula characters to prevent CSV/Excel injection."""
    if isinstance(value, str):
        value = value.strip()
        # Remove BOM, zero-width chars, etc.
        value = value.replace("\u200b", "").replace("\ufeff", "").replace("\u200c", "").replace("\u200d", "")
        if _FORMULA_RE.match(value):
            value = "'" + value  # Excel will display without formula execution
    return value


# ---------------------------------------------------------------------------
# Template Generation
# ---------------------------------------------------------------------------


def generate_sample_template() -> bytes:
    """Generate a plain sample .xlsx template.

    Produces a workbook with two sheets:
    - `INSTRUCTIONS`: short, clear steps (delete before importing)
    - `TEMPLATE`: headers and one example row using DD/MM/YYYY
    No colors or fancy styling to keep the file simple for non-technical users.
    """
    wb = Workbook()

    # Single-sheet TEMPLATE layout:
    # Row 1: headers (required columns first, then optional)
    # Row 2: single example data row
    # Row 3+: concise instructions that MUST be deleted before re-uploading
    ws = wb.active
    ws.title = "TEMPLATE"

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 14)

    # Example row (simple, no styling). Dates shown in DD/MM/YYYY as requested.
    for col_idx, val in enumerate(EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Concise instructions placed directly below example row.
    # Users MUST delete these instruction rows (and any example rows) before importing.
    ins_row = 3
    ws.cell(row=ins_row, column=1, value="IMPORTANT - READ BEFORE IMPORTING")
    ws.cell(row=ins_row + 1, column=1, value="1. DELETE ALL INSTRUCTION ROWS and any example rows BEFORE uploading this file.")
    ws.cell(row=ins_row + 2, column=1, value="2. Required columns (first): Name, Roll_Number, Registration_Number, Class.")
    ws.cell(row=ins_row + 3, column=1, value="3. Optional columns: Section, Father_Name, Father_CNIC, Gender, Date_of_Birth, Parent_Contact, Address, Admission_Date, Image_Name.")
    ws.cell(row=ins_row + 4, column=1, value="4. Date format must be DD/MM/YYYY (e.g., 31/12/2015).")
    ws.cell(row=ins_row + 5, column=1, value="5. Uniqueness: ONLY Registration_Number must be unique. Files with duplicate Registration_Number will be rejected.")
    ws.cell(row=ins_row + 6, column=1, value="6. If including images, upload a ZIP; filenames must exactly match Image_Name entries.")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


# ---------------------------------------------------------------------------
# Parsing & Validation
# ---------------------------------------------------------------------------


def _normalize_gender(raw: str) -> str:
    """Normalize gender value; raise ValueError if invalid."""
    key = raw.strip().lower()
    if key in GENDER_MAP:
        return GENDER_MAP[key]
    if not key:
        return ""
    raise ValueError(f"Invalid gender value: '{raw}'. Expected Male/Female/Other")


def _validate_date(raw: str, col_name: str) -> str:
    """Validate date strings.

    Accepts either DD/MM/YYYY (preferred in templates) or YYYY-MM-DD and
    returns a normalized YYYY-MM-DD string used internally.
    """
    if not raw:
        return ""
    raw = raw.strip()

    # Try DD/MM/YYYY first (user-facing template format), then YYYY-MM-DD
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    raise ValueError(f"Invalid date format for {col_name}: '{raw}'. Expected DD/MM/YYYY (e.g., 31/12/2015) or YYYY-MM-DD")


def _build_col_map(header_row: tuple) -> Dict[str, int]:
    """Build a case-insensitive header → column-index map with alias support."""
    col_map: Dict[str, int] = {}
    cleaned_headers = []
    
    for idx, h in enumerate(header_row):
        name = str(h).strip() if h is not None else ""
        cleaned_headers.append(name)
        # Normalize to match template column names (case-insensitive, underscores)
        key = name.replace(" ", "_").lower()
        
        # Check if this header matches any known alias
        matched = False
        for canonical_name, aliases in COLUMN_ALIASES.items():
            if key in aliases:
                col_map[canonical_name] = idx
                matched = True
                break
        
        # If no alias match, store the raw normalized key
        if not matched:
            col_map[key] = idx
    
    return col_map


def _check_required_columns(col_map: Dict[str, int]) -> Tuple[Optional[str], List[str]]:
    """
    Validate that all REQUIRED columns are present in the uploaded file.
    Extra columns are ignored. Column order doesn't matter.
    
    Returns:
        Tuple of (error_message or None, list of missing columns)
    """
    found_keys = set(col_map.keys())
    missing = REQUIRED_COLUMNS - found_keys
    if missing:
        # Create friendly display names
        display_map = {
            "name": "Name",
            "roll_number": "Roll Number",
            "registration_number": "Registration Number",
            "class": "Class",
        }
        missing_display = ", ".join([display_map.get(c, c.replace("_", " ").title()) for c in sorted(missing)])
        return f"Your file is missing the following required columns: {missing_display}. Please add these columns and try again.", list(missing)
    return None, []


def parse_and_validate_rows(xlsx_bytes: bytes) -> Dict[str, Any]:
    """
    Parse xlsx bytes.  Return:
    {
        "total_rows": int,
        "valid_rows": [...],
        "error_rows": [...],   # {row, column, value, reason}
        "duplicate_rows": [...],
        "duplicate_ids": set,
    }
    No DB access here — pure data validation.
    """
    wb = openpyxl.load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"total_rows": 0, "valid_rows": [], "error_rows": [{"row": 0, "column": "-", "value": "-", "reason": "Empty workbook"}], "duplicate_rows": [], "duplicate_ids": set()}

    col_map = _build_col_map(header_row)

    # Flexible template validation — check required columns only
    mismatch_err, missing_cols = _check_required_columns(col_map)
    if mismatch_err:
        return {
            "total_rows": 0, 
            "valid_rows": [], 
            "error_rows": [{"row": 0, "column": "-", "value": "-", "reason": mismatch_err}], 
            "duplicate_rows": [], 
            "duplicate_ids": set(),
            "missing_columns": missing_cols,
            "column_mapping_used": {k: v for k, v in col_map.items()},
        }

    valid_rows: List[Dict[str, Any]] = []
    error_rows: List[Dict[str, Any]] = []
    duplicate_rows: List[Dict[str, Any]] = []
    seen_reg_numbers: set = set()

    def _cell(row_data: tuple, key: str) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row_data) or row_data[idx] is None:
            return ""
        return str(_sanitize_cell(row_data[idx])).strip()

    for row_num, row_data in enumerate(rows_iter, start=2):
        # Skip completely empty rows
        if all(c is None or str(c).strip() == "" for c in row_data):
            continue

        row_errors: List[Dict[str, Any]] = []

        # Get values using new column names
        full_name = _cell(row_data, "name")
        roll_number = _cell(row_data, "roll_number")
        registration_number = _cell(row_data, "registration_number")
        class_val = _cell(row_data, "class")
        section = _cell(row_data, "section")  # Now optional
        gender_raw = _cell(row_data, "gender")
        dob_raw = _cell(row_data, "date_of_birth")
        father_name = _cell(row_data, "father_name")
        father_cnic = _cell(row_data, "father_cnic")
        parent_contact = _cell(row_data, "parent_contact")
        address = _cell(row_data, "address")
        admission_date_raw = _cell(row_data, "admission_date")
        image_name = _cell(row_data, "image_name")

        # Required field checks (only 4 required now)
        if not full_name:
            row_errors.append({"row": row_num, "column": "Name", "value": "", "reason": "Name is required", "status": "skipped"})
        if not roll_number:
            row_errors.append({"row": row_num, "column": "Roll_Number", "value": "", "reason": "Roll Number is required", "status": "skipped"})
        if not registration_number:
            row_errors.append({"row": row_num, "column": "Registration_Number", "value": "", "reason": "Registration Number is required", "status": "skipped"})
        if not class_val:
            row_errors.append({"row": row_num, "column": "Class", "value": "", "reason": "Class is required", "status": "skipped"})

        # Gender normalization
        gender = ""
        if gender_raw:
            try:
                gender = _normalize_gender(gender_raw)
            except ValueError as e:
                row_errors.append({"row": row_num, "column": "Gender", "value": gender_raw, "reason": str(e)})

        # Date validation
        dob = ""
        if dob_raw:
            try:
                dob = _validate_date(dob_raw, "Date_of_Birth")
            except ValueError as e:
                row_errors.append({"row": row_num, "column": "Date_of_Birth", "value": dob_raw, "reason": str(e)})

        admission_date = ""
        if admission_date_raw:
            try:
                admission_date = _validate_date(admission_date_raw, "Admission_Date")
            except ValueError as e:
                row_errors.append({"row": row_num, "column": "Admission_Date", "value": admission_date_raw, "reason": str(e)})

        # In-file duplicate detection: ONLY Registration_Number is required to be unique
        is_dup = False
        if registration_number and registration_number in seen_reg_numbers:
            duplicate_rows.append({
                "row": row_num,
                "column": "Registration_Number",
                "value": registration_number,
                "reason": "This Registration Number appears more than once in your file",
                "status": "skipped",
            })
            is_dup = True

        if registration_number:
            seen_reg_numbers.add(registration_number)

        if row_errors:
            error_rows.extend(row_errors)
            continue

        if is_dup:
            continue

        # Build normalized row with all data (missing fields left blank)
        valid_rows.append({
            "row_num": row_num,
            "registration_number": registration_number,
            "full_name": full_name,
            "roll_number": roll_number,
            "class_id": class_val,
            "section": section or "",  # Optional now
            "gender": gender or "",
            "date_of_birth": dob or "",
            "father_name": father_name or "",
            "father_cnic": father_cnic or "",
            "parent_contact": parent_contact or "",
            "address": address or "",
            "admission_date": admission_date or datetime.utcnow().strftime("%Y-%m-%d"),
            "image_name": image_name,
        })

    wb.close()

    return {
        "total_rows": len(valid_rows) + len(error_rows) + len(duplicate_rows),
        "valid_rows": valid_rows,
        "error_rows": error_rows,
        "duplicate_rows": duplicate_rows,
        "duplicate_ids": seen_reg_numbers,
    }


def check_db_duplicates(valid_rows: List[Dict], db, school_id: str = None) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    Check valid rows against database for duplicates by registration_number.
    Also ensures class exists or will be created.
    """
    clean: List[Dict] = []
    db_dups: List[Dict] = []
    existing_for_update: List[Dict] = []

    for row in valid_rows:
        reg_number = row["registration_number"]
        generated_student_id = generate_imported_student_id(reg_number)

        # Check if registration_number already exists (with school_id isolation)
        query = {"$or": [
            {"student_id": generated_student_id},
            {"registration_number": reg_number}
        ]}
        if school_id:
            query["school_id"] = school_id
        
        existing_student = db.students.find_one(query)

        if existing_student:
            db_dups.append({
                "row": row["row_num"], 
                "column": "Registration_Number", 
                "value": reg_number, 
                "reason": f"A student with this Registration Number already exists in the system",
                "status": "skipped"
            })
            existing_for_update.append({**row, "_existing_id": str(existing_student["_id"]), "student_id": generated_student_id})
        else:
            row["student_id"] = generated_student_id
            row["admission_year"] = 0  # For imported students
            clean.append(row)

    return clean, db_dups, existing_for_update

    return clean, db_dups, existing_for_update


# ---------------------------------------------------------------------------
# Import Execution (transactional)
# ---------------------------------------------------------------------------


def build_student_doc(row: Dict) -> Dict:
    """Convert a validated row dict into the document shape expected by create_student."""
    now = datetime.utcnow()
    
    # Determine data completeness status
    required_optional_fields = ["section", "father_name", "father_cnic", "gender", "date_of_birth", "parent_contact", "address"]
    missing_fields = [f for f in required_optional_fields if not row.get(f)]
    data_status = "complete" if not missing_fields else "incomplete"
    
    return {
        "student_id": row.get("student_id", ""),
        "registration_number": row.get("registration_number", ""),
        "full_name": row["full_name"],
        "roll_number": row["roll_number"],
        "class_id": row["class_id"],
        "section": row.get("section", ""),
        "gender": row.get("gender", ""),
        "date_of_birth": row.get("date_of_birth", ""),
        "admission_date": row.get("admission_date", "") or now.strftime("%Y-%m-%d"),
        "guardian_info": {
            "father_name": row.get("father_name", ""),
            "parent_cnic": row.get("father_cnic", ""),
            "guardian_contact": row.get("parent_contact", ""),
            "address": row.get("address", ""),
        },
        "contact_info": {
            "phone": row.get("parent_contact", ""),
        },
        "subjects": [],
        "assigned_teacher_ids": [],
        "status": "active",
        "data_status": data_status,
        "missing_fields": missing_fields,
        "academic_year": f"{now.year}-{now.year + 1}",
        "created_at": now,
        "updated_at": now,
    }


def execute_import_transaction(
    rows_to_insert: List[Dict],
    rows_to_update: List[Dict],
    duplicate_action: str,
    db,
) -> Tuple[int, int, List[Dict]]:
    """
    Execute the import inside a MongoDB client session / transaction where supported.
    Returns (success_count, fail_count, errors).
    All-or-nothing: if any write fails, the entire batch is rolled back.
    """
    docs_to_insert = [build_student_doc(r) for r in rows_to_insert]

    update_docs = []
    if duplicate_action == "update":
        for r in rows_to_update:
            doc = build_student_doc(r)
            doc.pop("created_at", None)
            update_docs.append((r["_existing_id"], doc))

    success = 0
    errors: List[Dict] = []

    # Try using a session/transaction (requires replica set)
    client = db.client
    try:
        with client.start_session() as session:
            with session.start_transaction():
                if docs_to_insert:
                    db.students.insert_many(docs_to_insert, session=session)
                    success += len(docs_to_insert)

                for eid, doc in update_docs:
                    from bson.objectid import ObjectId
                    db.students.update_one(
                        {"_id": ObjectId(eid)},
                        {"$set": doc},
                        session=session,
                    )
                    success += 1
    except Exception:
        # Standalone MongoDB — no transaction support; fall back to ordered bulk
        success = 0
        errors = []
        try:
            if docs_to_insert:
                db.students.insert_many(docs_to_insert, ordered=True)
                success += len(docs_to_insert)

            for eid, doc in update_docs:
                from bson.objectid import ObjectId
                db.students.update_one({"_id": ObjectId(eid)}, {"$set": doc})
                success += 1
        except Exception as exc:
            # If bulk insert partially fails, we can't easily roll back in standalone.
            # Delete any docs just inserted in this batch to honour all-or-nothing.
            inserted_ids = [d["student_id"] for d in docs_to_insert]
            if inserted_ids:
                db.students.delete_many({"student_id": {"$in": inserted_ids}})
            return 0, len(docs_to_insert) + len(update_docs), [{"row": 0, "column": "-", "value": "-", "reason": f"Transaction failed: {str(exc)}"}]

    return success, 0, errors


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

# Export columns include all template columns plus a Status column
EXPORT_COLUMNS = TEMPLATE_COLUMNS + ["Data_Status"]


def export_students_xlsx(students: List[Dict]) -> bytes:
    """Export students list to xlsx bytes with all columns plus data status."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_required = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_fill_optional = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill_status = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        if col_name == "Data_Status":
            cell.fill = header_fill_status
        elif col_name.lower().replace(" ", "_") in REQUIRED_COLUMNS:
            cell.fill = header_fill_required
        else:
            cell.fill = header_fill_optional
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 6, 16)

    # Status cell styling
    complete_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    incomplete_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    for row_idx, s in enumerate(students, start=2):
        guardian = s.get("guardian_info") or {}
        contact = s.get("contact_info") or {}
        
        # Determine data status
        data_status = s.get("data_status", "")
        if not data_status:
            # Calculate if not stored
            optional_fields = ["section", "father_name", "father_cnic", "gender", "date_of_birth", "parent_contact", "address"]
            missing = []
            if not s.get("section"): missing.append("section")
            if not guardian.get("father_name"): missing.append("father_name")
            if not guardian.get("parent_cnic"): missing.append("father_cnic")
            if not s.get("gender"): missing.append("gender")
            if not s.get("date_of_birth"): missing.append("date_of_birth")
            if not guardian.get("guardian_contact") and not contact.get("phone"): missing.append("parent_contact")
            if not guardian.get("address") and not s.get("address"): missing.append("address")
            data_status = "Complete" if not missing else "Incomplete"
        else:
            data_status = "Complete" if data_status == "complete" else "Incomplete"
        
        row_data = [
            _sanitize_cell(s.get("full_name", "")),
            _sanitize_cell(s.get("roll_number", "")),
            _sanitize_cell(s.get("registration_number", "") or s.get("student_id", "")),
            _sanitize_cell(s.get("class_id", "") or s.get("class", "")),
            _sanitize_cell(s.get("section", "")),
            _sanitize_cell(guardian.get("father_name", "")),
            _sanitize_cell(guardian.get("parent_cnic", "")),
            _sanitize_cell(s.get("gender", "")),
            _sanitize_cell(s.get("date_of_birth", "")),
            _sanitize_cell(guardian.get("guardian_contact", "") or contact.get("phone", "")),
            _sanitize_cell(guardian.get("address", "") or s.get("address", "")),
            _sanitize_cell(s.get("admission_date", "")),
            _sanitize_cell(s.get("image_name", "") or ""),
            data_status,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            # Color the status column
            if col_idx == len(EXPORT_COLUMNS):
                cell.fill = complete_fill if val == "Complete" else incomplete_fill

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


# ---------------------------------------------------------------------------
# Error Report
# ---------------------------------------------------------------------------


def generate_error_report(errors: List[Dict]) -> bytes:
    """Generate students_import_errors.xlsx from error list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Errors"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Row", "Column", "Provided Value", "Error Reason"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 25

    for row_idx, err in enumerate(errors, start=2):
        ws.cell(row=row_idx, column=1, value=err.get("row", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=err.get("column", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=_sanitize_cell(err.get("value", ""))).border = thin_border
        ws.cell(row=row_idx, column=4, value=err.get("reason", "")).border = thin_border

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
