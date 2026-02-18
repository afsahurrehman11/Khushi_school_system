from app.database import get_db
from datetime import datetime
from typing import Optional
from bson.objectid import ObjectId
from io import BytesIO
from typing import List, Tuple
from app.utils.student_id_utils import generate_student_id, validate_student_id_uniqueness
import logging
import zipfile

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore[import]
    from openpyxl import Workbook  # type: ignore[import]
except Exception:
    openpyxl = None

# ================= Student Operations =================

def create_student(student_data: dict) -> Optional[dict]:
    """Create a new student with schoolId isolation"""
    try:
        school_id = student_data.get("school_id")
        if not school_id:
            logger.error(f"❌ Cannot create student without schoolId")
            return None
            
        logger.info(f"[SCHOOL:{school_id}] 🎓 Creating student: {student_data.get('full_name', 'Unknown')}")
        db = get_db()

        # Set admission_year if not provided
        if "admission_year" not in student_data:
            student_data["admission_year"] = datetime.utcnow().year
            logger.debug(f"[SCHOOL:{school_id}] 📅 Set admission_year to {student_data['admission_year']}")

        # Generate student_id if not provided (manual creation)
        if "student_id" not in student_data:
            student_data["student_id"] = generate_student_id(student_data["admission_year"], db.students, school_id)
            logger.info(f"[SCHOOL:{school_id}] 🆔 Generated student_id: {student_data['student_id']}")

        # Validate uniqueness within school
        if not validate_student_id_uniqueness(student_data["student_id"], db.students, school_id):
            logger.error(f"[SCHOOL:{school_id}] ❌ Student ID already exists: {student_data['student_id']}")
            return None

        # Ensure email uniqueness constraint won't fail for missing emails.
        # The database has a unique index on (school_id, email). MongoDB treats null/None
        # as the same value, so multiple documents without email will conflict. To avoid
        # this, set a unique placeholder email derived from the generated student_id
        # when email is not provided.
        email_val = student_data.get('email')
        if not email_val:
            placeholder = f"{student_data['student_id']}@no-email.{school_id}.local"
            student_data['email'] = placeholder
            logger.debug(f"[SCHOOL:{school_id}] ℹ️ No email provided; set placeholder email: {placeholder}")

        student_data["created_at"] = datetime.utcnow()
        student_data["updated_at"] = datetime.utcnow()

        result = db.students.insert_one(student_data)
        student_data["id"] = str(result.inserted_id)
        logger.info(f"[SCHOOL:{school_id}] ✅ Student created successfully: {student_data['student_id']} - {student_data['full_name']}")
        return student_data
    except Exception as e:
        logger.error(f"❌ Failed to create student: {str(e)}")
        return None

def get_all_students(filters: dict = None, school_id: str = None) -> list:
    """Get all students with optional filters and schoolId isolation"""
    db = get_db()
    query = filters or {}
    
    # Enforce schoolId filtering
    if school_id:
        query["school_id"] = school_id
        logger.info(f"[SCHOOL:{school_id}] 📋 Fetching students")
    
    students = list(db.students.find(query))
    for student in students:
        # ensure id string
        student["id"] = str(student["_id"])
        # normalize missing fields so Pydantic response models don't fail
        student.setdefault('full_name', 'Unnamed Student')
        student.setdefault('gender', 'Not specified')
        student.setdefault('date_of_birth', datetime.utcnow().strftime('%Y-%m-%d'))
        student.setdefault('admission_date', datetime.utcnow().strftime('%Y-%m-%d'))
        student.setdefault('admission_year', datetime.utcnow().year)
        student.setdefault('roll_number', '')
        student.setdefault('academic_year', f"{datetime.utcnow().year}-{datetime.utcnow().year+1}")
        student.setdefault('subjects', student.get('subjects', []))
        student.setdefault('assigned_teacher_ids', student.get('assigned_teacher_ids', []))
        # ensure timestamps exist
        if 'created_at' not in student:
            student['created_at'] = datetime.utcnow()
        if 'updated_at' not in student:
            student['updated_at'] = datetime.utcnow()
    
    logger.info(f"[SCHOOL:{school_id}] ✅ Retrieved {len(students)} students") if school_id else None
    return students

def get_student_by_id(student_id: str, school_id: str = None) -> Optional[dict]:
    """Get student by ID with optional schoolId isolation"""
    db = get_db()
    try:
        query = {"_id": ObjectId(student_id)}
        if school_id:
            query["school_id"] = school_id
        
        student = db.students.find_one(query)
        if student:
            student["id"] = str(student["_id"])
            logger.info(f"[SCHOOL:{school_id or 'N/A'}] 🔍 Retrieved student: {student_id}") if school_id else None
        return student
    except:
        return None

def get_student_by_student_id(student_id: str, school_id: str = None) -> Optional[dict]:
    """Get student by student_id field with optional schoolId isolation"""
    db = get_db()
    query = {"student_id": student_id}
    if school_id:
        query["school_id"] = school_id
    
    student = db.students.find_one(query)
    if student:
        student["id"] = str(student["_id"])
        logger.info(f"[SCHOOL:{school_id or 'N/A'}] 🔍 Retrieved student by student_id: {student_id}") if school_id else None
    return student

def update_student(student_id: str, school_id: str = None, **kwargs) -> Optional[dict]:
    """Update student with schoolId isolation"""
    db = get_db()
    try:
        query = {"_id": ObjectId(student_id)}
        if school_id:
            query["school_id"] = school_id
        
        kwargs["updated_at"] = datetime.utcnow()
        result = db.students.find_one_and_update(
            query,
            {"$set": kwargs},
            return_document=True
        )
        if result:
            result["id"] = str(result["_id"])
            logger.info(f"[SCHOOL:{school_id or 'N/A'}] ✅ Student updated: {student_id}") if school_id else None
        return result
    except:
        return None

def delete_student(student_id: str, school_id: str = None) -> bool:
    """Delete student with schoolId isolation"""
    db = get_db()
    try:
        query = {"_id": ObjectId(student_id)}
        if school_id:
            query["school_id"] = school_id
        
        result = db.students.delete_one(query)
        if result.deleted_count > 0:
            logger.info(f"[SCHOOL:{school_id or 'N/A'}] ✅ Student deleted: {student_id}") if school_id else None
        return result.deleted_count > 0
    except:
        return False


# ---------------- Import / Export Helpers ----------------
def import_students_from_workbook_bytes(xlsx_bytes: bytes, class_id: str) -> Tuple[List[dict], List[dict]]:
    """Parse xlsx bytes, create students, return (created, errors)

    Expected columns (case-insensitive): name, father_name, parent_cnic, registration_ID, fathers_NIC, section, subjects
    """
    if openpyxl is None:
        raise RuntimeError('openpyxl not installed')

    wb = openpyxl.load_workbook(filename=BytesIO(xlsx_bytes), read_only=True)
    ws = wb.active

    created = []
    errors = []

    # read header
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return created, [{'row': 0, 'error': 'Empty workbook'}]

    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    # map expected columns
    col_map = {}
    for idx, col in enumerate(header):
        key = col.strip().lower()
        col_map[key] = idx

    required_keys = ['name', 'father_name', 'parent_cnic', 'registration_id']
    for rk in required_keys:
        if rk not in col_map:
            return created, [{'row': 0, 'error': f'Missing required column: {rk}'}]

    for r_idx, row in enumerate(rows[1:], start=2):
        try:
            # read values safely
            def val(k):
                i = col_map.get(k)
                return str(row[i]).strip() if i is not None and row[i] is not None else ''

            name = val('name')
            father_name = val('father_name')
            parent_cnic = val('parent_cnic')
            older_reg = val('registration_id')
            fathers_nic = val('fathers_nic')
            section = val('section') or ''
            subjects_raw = val('subjects')
            subjects = [s.strip() for s in subjects_raw.split(',')] if subjects_raw else []

            if not name or not older_reg:
                errors.append({'row': r_idx, 'error': 'Missing required name or registration_ID'})
                continue

            if not parent_cnic:
                errors.append({'row': r_idx, 'error': 'Missing required parent_cnic'})
                continue

            student_id = f"0000-{older_reg}"

            # skip duplicates
            if get_student_by_student_id(student_id):
                errors.append({'row': r_idx, 'error': 'Duplicate registration_ID, skipped', 'registration_id': student_id})
                continue

            student_doc = {
                'student_id': student_id,
                'full_name': name,
                'class_id': class_id,
                'section': section,
                'subjects': subjects,
                'guardian_info': {
                    'parent_cnic': parent_cnic,
                    'father_name': father_name,
                    'fathers_NIC': fathers_nic,
                }
            }

            res = create_student(student_doc)
            if res is None:
                errors.append({'row': r_idx, 'error': 'Failed to create (duplicate?)', 'registration_id': student_id})
            else:
                created.append(res)
        except Exception as e:
            errors.append({'row': r_idx, 'error': str(e)})

    return created, errors


def parse_students_from_workbook_bytes(xlsx_bytes: bytes, max_preview: int = 10) -> List[dict]:
    """Parse xlsx bytes and return up to max_preview parsed rows (no DB writes)."""
    if openpyxl is None:
        raise RuntimeError('openpyxl not installed')

    wb = openpyxl.load_workbook(filename=BytesIO(xlsx_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    col_map = {col.strip().lower(): idx for idx, col in enumerate(header)}

    parsed = []
    for r_idx, row in enumerate(rows[1:1+max_preview], start=2):
        def val(k):
            i = col_map.get(k)
            return str(row[i]).strip() if i is not None and row[i] is not None else ''

        parsed.append({
            'row': r_idx,
            'name': val('name'),
            'father_name': val('father_name'),
            'parent_cnic': val('parent_cnic'),
            'registration_ID': val('registration_id'),
            'fathers_NIC': val('fathers_nic'),
            'section': val('section'),
            'subjects': [s.strip() for s in (val('subjects') or '').split(',') if s.strip()]
        })
    return parsed


def export_students_to_workbook_bytes(filters: dict = None) -> bytes:
    """Export students matching filters to xlsx bytes"""
    if openpyxl is None:
        raise RuntimeError('openpyxl not installed')

    students = get_all_students(filters or {})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Students'

    header = ['name', 'father_name', 'parent_cnic', 'registration_ID', 'fathers_NIC', 'class', 'section', 'subjects']
    ws.append(header)

    for s in students:
        name = s.get('full_name') or s.get('name') or ''
        guardian = s.get('guardian_info') or {}
        father_name = guardian.get('father_name', '')
        parent_cnic = guardian.get('parent_cnic', '')
        fathers_nic = guardian.get('fathers_NIC', '')
        reg = s.get('student_id') or ''
        class_name = s.get('class') or s.get('class_id') or ''
        section = s.get('section') or ''
        subjects = ','.join(s.get('subjects') or [])
        ws.append([name, father_name, parent_cnic, reg, fathers_nic, class_name, section, subjects])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def import_students_with_images(xlsx_bytes: bytes, zip_bytes: Optional[bytes], class_id: str) -> dict:
    """
    Import students from Excel with optional ZIP file containing images
    
    Args:
        xlsx_bytes: Excel file bytes
        zip_bytes: Optional ZIP file containing student images
        class_id: Class ID for students
        
    Returns:
        Dict with import summary
    """
    try:
        logger.info("📊 Starting bulk import with images")
        
        # Parse Excel
        created_students = []
        failed_students = []
        
        created, errors = import_students_from_workbook_bytes(xlsx_bytes, class_id)
        created_students.extend(created)
        failed_students.extend(errors)
        
        images_uploaded = 0
        image_failures = 0
        
        # Process images if ZIP provided
        if zip_bytes:
            logger.info("📦 Processing images from ZIP file")
            
            try:
                # Extract and process images
                with zipfile.ZipFile(BytesIO(zip_bytes), 'r') as zip_ref:
                    for file_info in zip_ref.filelist:
                        try:
                            filename = file_info.filename
                            
                            # Skip directories and hidden files
                            if filename.endswith('/') or filename.startswith('.'):
                                continue
                            
                            # Extract filename without extension
                            base_name = filename.split('/')[-1].split('.')[0]
                            
                            # Try to match with created students by registration_id or student_id
                            matching_student = None
                            for student in created_students:
                                student_id_val = student.get('student_id', '')
                                # Match by student_id or base filename
                                if base_name in student_id_val or student_id_val in base_name:
                                    matching_student = student
                                    break
                            
                            if matching_student:
                                # Read image
                                image_content = zip_ref.read(filename)
                                
                                # Lazy import to avoid circular dependency
                                from app.services.cloudinary_service import CloudinaryService
                                
                                student_id = str(matching_student.get('_id', ''))
                                upload_result = CloudinaryService.upload_image(
                                    image_content,
                                    filename,
                                    student_id
                                )
                                
                                if upload_result:
                                    # Update student with image
                                    db = get_db()
                                    db.students.update_one(
                                        {"_id": ObjectId(student_id)},
                                        {
                                            "$set": {
                                                "profile_image_url": upload_result["secure_url"],
                                                "profile_image_public_id": upload_result["public_id"],
                                                "image_uploaded_at": datetime.utcnow(),
                                                "embedding_status": "pending",
                                                "updated_at": datetime.utcnow()
                                            }
                                        }
                                    )
                                    images_uploaded += 1
                                    logger.info(f"✅ Image uploaded for {matching_student.get('student_id')}")
                                else:
                                    image_failures += 1
                                    logger.warning(f"⚠️ Failed to upload image for {matching_student.get('student_id')}")
                        except Exception as e:
                            image_failures += 1
                            logger.error(f"Error processing image {filename}: {str(e)}")
            
            except zipfile.BadZipFile:
                logger.error("Invalid ZIP file provided")
                failed_students.append({
                    'row': 'ZIP',
                    'error': 'Invalid or corrupted ZIP file'
                })
        
        logger.info(f"✅ Bulk import completed")
        
        return {
            "success": True,
            "summary": {
                "total_students_created": len(created_students),
                "failed_students": len(failed_students),
                "images_uploaded": images_uploaded,
                "image_failures": image_failures
            },
            "created": created_students,
            "errors": failed_students
        }
    
    except Exception as e:
        logger.error(f"❌ Bulk import failed: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }