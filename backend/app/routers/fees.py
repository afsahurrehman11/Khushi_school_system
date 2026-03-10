from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Optional
import logging
from io import BytesIO
from app.models.fee import FeeCreate, FeeInDB, FeeUpdate, FeeGenerate
from app.services.fee import (
    get_all_fees, get_fee_by_id, create_fee, update_fee, delete_fee, get_fees_by_student
)
from app.services.student import get_all_students
from app.dependencies.auth import check_permission
from app.database import get_db
from typing import Dict
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

def convert_objectids(obj):
    """Recursively convert ObjectId to string in dict/list"""
    if isinstance(obj, dict):
        return {k: convert_objectids(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_objectids(item) for item in obj]
    elif isinstance(obj, ObjectId):
        return str(obj)
    else:
        return obj

router = APIRouter()

@router.get("/fees")
async def get_fees(
    student_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    sort_by: str = "created_at",
    sort_dir: str = "desc",
    current_user: dict = Depends(check_permission("fees.view"))
):
    """Get fees with optional filters and pagination. Returns {count, page, page_size, fees}"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id or 'All'}] [ADMIN:{admin_email}] Fetching fees")
    
    try:
        db = get_db()
        query = {"school_id": school_id} if school_id else {}
        if student_id:
            query["student_id"] = student_id
        if status:
            query["status"] = status

        # sanitize paging
        if page < 1:
            page = 1
        if page_size < 1 or page_size > 500:
            page_size = 20

        total = db.fees.count_documents(query)
        sd = -1 if sort_dir.lower() == "desc" else 1
        cursor = db.fees.find(query).sort(sort_by, sd).skip((page - 1) * page_size).limit(page_size)
        fees = list(cursor)
        fees = [convert_objectids(fee) for fee in fees]
        for f in fees:
            f["id"] = f.pop("_id", None)

        logger.info(f"[SCHOOL:{school_id or 'All'}] ✅ Retrieved {len(fees)} fees")
        return {
            "count": total,
            "page": page,
            "page_size": page_size,
            "fees": fees,
        }
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id or 'All'}] ❌ Failed to fetch fees: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch fees")





# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTANT: /fees/export MUST be defined BEFORE /fees/{fee_id} to avoid route conflicts!
# FastAPI matches routes in order, so /fees/export would match /fees/{fee_id} with fee_id="export"
# ═══════════════════════════════════════════════════════════════════════════════

# Import PDF export function at the top of routes section
from app.routers.fees_export_pdf import export_fees_by_status_pdf

@router.get("/fees/export")
async def export_fees_by_status(
    class_id: str,
    status: str,
    section: str = None,
    current_user: dict = Depends(check_permission("fees.view"))
):
    """
    Export fee report by status (paid/partial/unpaid) as PDF.
    
    This endpoint MUST be defined before /fees/{fee_id} to avoid route conflicts.
    """
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    
    # Extensive entry logging
    print(f"\n{'═'*80}")
    print(f"[FEES_EXPORT] ✅ ROUTE HIT: /fees/export")
    print(f"[FEES_EXPORT] class_id={class_id}, status={status}, section={section}")
    print(f"[FEES_EXPORT] school_id={school_id}, admin={admin_email}")
    print(f"[FEES_EXPORT] Delegating to PDF export module...")
    print(f"{'═'*80}")
    
    logger.info(f"[SCHOOL:{school_id}] [ADMIN:{admin_email}] [FEES_EXPORT] Exporting {status} fees for class {class_id} section {section}")
    
    try:
        result = await export_fees_by_status_pdf(class_id, status, section, current_user)
        print(f"[FEES_EXPORT] ✅ PDF export completed successfully")
        logger.info(f"[SCHOOL:{school_id}] [FEES_EXPORT] ✅ PDF export completed")
        return result
    except HTTPException as he:
        print(f"[FEES_EXPORT] ❌ HTTPException: {he.status_code} - {he.detail}")
        logger.error(f"[SCHOOL:{school_id}] [FEES_EXPORT] ❌ HTTPException: {he.detail}")
        raise
    except Exception as e:
        print(f"[FEES_EXPORT] ❌ EXCEPTION: {str(e)}")
        import traceback
        traceback.print_exc()
        logger.error(f"[SCHOOL:{school_id}] [FEES_EXPORT] ❌ Exception: {str(e)}")
        raise HTTPException(status_code=500, detail=f"PDF export failed: {str(e)}")

@router.get("/fees/search", response_model=List[FeeInDB])
async def search_fees(
    student_name: Optional[str] = None,
    class_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(check_permission("fees.view"))
):
    """Search fees by student name, class, and/or status"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id or 'All'}] [ADMIN:{admin_email}] Searching fees")
    
    try:
        filters: Dict = {"school_id": school_id} if school_id else {}
        if class_id:
            filters["class_id"] = class_id
        if status:
            filters["status"] = status

        # If student_name provided, find matching students then filter by their student_id
        if student_name:
            student_filters = {"full_name": {"$regex": student_name, "$options": "i"}}
            if school_id:
                student_filters["school_id"] = school_id
            students = get_all_students(student_filters)
            if students:
                student_ids = [s.get("student_id") for s in students if s.get("student_id")]
                if student_ids:
                    filters["student_id"] = {"$in": student_ids}
                else:
                    logger.warning(f"[SCHOOL:{school_id or 'All'}] No matching student IDs found")
                    return []
            else:
                logger.warning(f"[SCHOOL:{school_id or 'All'}] No matching students found")
                return []

        fees = get_all_fees(filters, school_id=school_id)
        logger.info(f"[SCHOOL:{school_id or 'All'}] ✅ Search returned {len(fees)} fees")
        return fees
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id or 'All'}] ❌ Failed to search fees: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to search fees")

# NOTE: /fees/{fee_id} MUST come AFTER /fees/export and /fees/search to avoid route conflicts
@router.get("/fees/{fee_id}", response_model=FeeInDB)
async def get_fee(
    fee_id: str,
    current_user: dict = Depends(check_permission("fees.view"))
):
    """Get fee by ID"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id or 'All'}] [ADMIN:{admin_email}] Fetching fee {fee_id}")
    
    try:
        fee = get_fee_by_id(fee_id, school_id=school_id)
        if not fee:
            logger.warning(f"[SCHOOL:{school_id or 'All'}] Fee {fee_id} not found")
            raise HTTPException(status_code=404, detail="Fee not found")
        logger.info(f"[SCHOOL:{school_id or 'All'}] ✅ Fee {fee_id} found")
        return fee
    except HTTPException as he:
        # If nothing was found, return an empty Excel file instead of 404 so frontend can download a file
        if getattr(he, 'status_code', None) == 404:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Fees"
                headers = ["Name", "Roll Number", "Registration ID", "Class", "Section", "Total Fee", "Paid Amount", "Remaining", "Arrears", "Scholarship %", "Scholarship Amount", "Status"]
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                    ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 14)

                ws.cell(row=2, column=1, value=str(he.detail or "No records found"))

                bio = BytesIO()
                wb.save(bio)
                bio.seek(0)

                from datetime import datetime
                filename = f"fee_report_empty_{datetime.now().strftime('%Y%m%d')}.xlsx"

                return StreamingResponse(
                    bio,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )
            except Exception:
                # If generating the empty file fails, re-raise the original 404
                raise
        raise
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id or 'All'}] ❌ Failed to fetch fee: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch fee")

@router.get("/students/{student_id}/fees", response_model=List[FeeInDB])
async def get_student_fees(
    student_id: str,
    current_user: dict = Depends(check_permission("fees.view"))
):
    """Get all fees for a specific student"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id or 'All'}] [ADMIN:{admin_email}] Fetching fees for student {student_id}")
    
    try:
        fees = get_fees_by_student(student_id, school_id=school_id)
        logger.info(f"[SCHOOL:{school_id or 'All'}] ✅ Retrieved {len(fees)} fees for student {student_id}")
        return fees
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id or 'All'}] ❌ Failed to fetch student fees: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch student fees")

@router.post("/fees", response_model=FeeInDB)
async def create_new_fee(
    fee_data: FeeCreate,
    current_user: dict = Depends(check_permission("fees.manage"))
):
    """Create new fee"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id}] [ADMIN:{admin_email}] Creating fee")
    
    try:
        fee_dict = fee_data.dict()
        fee_dict["generated_by"] = current_user["id"]  # Set the user who created the fee
        fee = create_fee(fee_dict, school_id=school_id)
        if not fee:
            logger.warning(f"[SCHOOL:{school_id}] Fee creation failed")
            raise HTTPException(status_code=400, detail="Invalid fee data")
        logger.info(f"[SCHOOL:{school_id}] ✅ Fee {fee.get('_id')} created")
        return fee
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id}] ❌ Failed to create fee: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create fee")

@router.post("/fees/generate")
async def generate_fees(
    fee_data: FeeGenerate,
    current_user: dict = Depends(check_permission("fees.manage"))
):
    """Generate fees for multiple students"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id}] [ADMIN:{admin_email}] Generating fees for class")
    
    try:
        filters = {"class_id": fee_data.class_id}
        if school_id:
            filters["school_id"] = school_id
        students = get_all_students(filters) if hasattr(fee_data, 'class_id') else []

        created_fees = []
        for student in students:
            fee_dict = {
                "student_id": student["student_id"],
                "class_id": student["class_id"],
                "fee_type": fee_data.fee_type,
                "amount": fee_data.amount,
                "due_date": fee_data.due_date,
                "status": "pending",
                "generated_by": current_user["id"]
            }
            fee = create_fee(fee_dict, school_id=school_id)
            if fee:
                created_fees.append(fee)

        logger.info(f"[SCHOOL:{school_id}] ✅ Generated {len(created_fees)} fees")
        return {"count": len(created_fees), "fees": created_fees}
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id}] ❌ Failed to generate fees: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate fees")

@router.put("/fees/{fee_id}", response_model=FeeInDB)
async def update_existing_fee(
    fee_id: str,
    fee_data: FeeUpdate,
    current_user: dict = Depends(check_permission("fees.manage"))
):
    """Update fee"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id}] [ADMIN:{admin_email}] Updating fee {fee_id}")
    
    try:
        update_data = fee_data.dict(exclude_unset=True)
        fee = update_fee(fee_id, school_id=school_id, **update_data)
        if not fee:
            logger.warning(f"[SCHOOL:{school_id}] Fee {fee_id} not found")
            raise HTTPException(status_code=404, detail="Fee not found")
        logger.info(f"[SCHOOL:{school_id}] ✅ Fee {fee_id} updated")
        return fee
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id}] ❌ Failed to update fee: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update fee")

@router.delete("/fees/{fee_id}")
async def delete_existing_fee(
    fee_id: str,
    current_user: dict = Depends(check_permission("fees.manage"))
):
    """Delete fee"""
    school_id = current_user.get("school_id")
    admin_email = current_user.get("email")
    logger.info(f"[SCHOOL:{school_id}] [ADMIN:{admin_email}] Deleting fee {fee_id}")
    
    try:
        if not delete_fee(fee_id, school_id=school_id):
            logger.warning(f"[SCHOOL:{school_id}] Fee {fee_id} not found")
            raise HTTPException(status_code=404, detail="Fee not found")
        logger.info(f"[SCHOOL:{school_id}] ✅ Fee {fee_id} deleted")
        return {"message": "Fee deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[SCHOOL:{school_id}] ❌ Failed to delete fee: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete fee")